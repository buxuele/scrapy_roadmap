# -*- coding: utf-8 -*-
# author: fanchuang
# DateTime:2021/2/17 0017 18:28 
# contact: fanchuangwater@gmail.com
# about: python 操作 Excel

import os
import time
import shutil
from datetime import datetime
from openpyxl import load_workbook

""" 
使用 openpyxl 自动生成每日任务清单。

1. 自动生成每天的固定任务。随时可以添加临时任务。 
2. 最好是能定时执行，比如开机执行。
3. 写个文件来自动生成每日的todo任务。保存到桌面。格式最好是 excel.
4. 如果我已经有了一个文件模板，那么我直接在这个模板文件中新建一个工作表即可。
"""


# 数据表的名称
def make_name():
    a = datetime.date(datetime.now())
    return str(a)


def parse_text():
    raw_data = """
1. 半个小时。温习笔记，整理笔记
2. 半个小时。阅读文档，阅读源码
3. 2个爬虫。时间不限，但是要注重质量
4. 1个小时。写篇文章，发布文章
5. 1个小时。读技术类的书， 做好笔记
"""
    return raw_data.split('\n')


def generate_tasks():
    # 1. 准备工作，打开关闭等  始终更新同一个文件。
    excel_name = r'C:\Users\Administrator\Desktop\Small_Plan.xlsx'
    sheet_name = make_name()

    wb = load_workbook(excel_name)
    if sheet_name not in wb.sheetnames:
        working_sheet = wb.create_sheet(sheet_name, 0)  # 添加到开头，很机智啊。
    else:
        working_sheet = wb[sheet_name]

    data = parse_text()
    for index, things in enumerate(data):
        if things:
            print(things)
            order = things.split('. ')[0]
            working_sheet[f'A{index + 1}'] = order

            duration, work = things.split('. ')[1].split('。')
            working_sheet[f'B{index + 1}'] = duration
            working_sheet[f'C{index + 1}'] = work

    wb.save(excel_name)


if __name__ == '__main__':
    generate_tasks()

